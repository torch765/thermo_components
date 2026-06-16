"""Excel report exporter implemented with OpenPyXL."""

from pathlib import Path

from thermo_components.application.dto import ReportExportRequest


def get_excel_number_format(unit: str) -> str:
    """Return a sensible Excel number format for the given engineering unit."""
    if unit in {"Mol %", "Wt %"}:
        return "0.0000"
    if unit == "kg/m³":
        return "0.000"
    if unit in {
        "GJ/Nm³",
        "MMBtu/Nm³",
        "MMkcal/Nm³",
        "GJ/kg",
        "MMBtu/kg",
        "MMkcal/kg",
    }:
        return "0.0000"
    if unit:
        return "0.00"
    return "General"


class OpenPyxlReportExporter:
    """Write a prepared report projection to an Excel workbook."""

    def export(self, request: ReportExportRequest) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
            )
        except ImportError as exc:
            raise ImportError(
                "Excel export requires the 'openpyxl' package."
            ) from exc

        report_path = Path(request.report_path)
        report_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Results"
        worksheet.freeze_panes = "A5"

        thin_side = Side(style="thin", color="B7C9E2")
        table_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(italic=True, color="666666")
        section_font = Font(size=11, bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="5B9BD5")
        left_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        right_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        def style_row(row_index, max_col=4, fill=None, font=None):
            for col_index in range(1, max_col + 1):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment
                if fill is not None:
                    cell.fill = fill
                if font is not None:
                    cell.font = font

        def write_section_header(row_index, title):
            worksheet.cell(row=row_index, column=1, value=title)
            style_row(row_index, fill=section_fill, font=section_font)
            return row_index + 1

        def write_table_header(row_index, headers):
            for col_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(
                    row=row_index,
                    column=col_index,
                    value=header,
                )
                cell.font = header_font
                cell.fill = header_fill
                cell.border = table_border
                cell.alignment = left_alignment
            return row_index + 1

        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "THERMO CALCULATOR REPORT"
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = center_alignment
        worksheet.row_dimensions[1].height = 24

        worksheet.merge_cells("A2:D2")
        worksheet["A2"] = (
            f"Exported: {request.exported_at.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        worksheet["A2"].font = subtitle_font
        worksheet["A2"].alignment = center_alignment

        row_index = 4
        row_index = write_section_header(row_index, "Conditions")
        row_index = write_table_header(row_index, ["Setting", "Value"])
        for condition in request.conditions:
            worksheet.cell(row=row_index, column=1, value=condition.setting)
            value_cell = worksheet.cell(
                row=row_index,
                column=2,
                value=condition.value,
            )
            for col_index in range(1, 3):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment
            value_cell.alignment = left_alignment
            row_index += 1

        row_index += 1
        row_index = write_section_header(row_index, "Input Composition")
        row_index = write_table_header(
            row_index,
            ["Component", "Mol %", "Wt %"],
        )
        for composition_row in request.input_rows:
            worksheet.cell(
                row=row_index,
                column=1,
                value=composition_row.component,
            )
            for column_index, key, value in (
                (2, "Mol %", composition_row.mole_percent),
                (3, "Wt %", composition_row.weight_percent),
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value if value != "" else None,
                )
                cell.border = table_border
                if isinstance(value, (int, float)):
                    cell.alignment = right_alignment
                    cell.number_format = get_excel_number_format(key)
                else:
                    cell.alignment = left_alignment
            worksheet.cell(row=row_index, column=1).border = table_border
            worksheet.cell(row=row_index, column=1).alignment = left_alignment
            row_index += 1

        if request.projection.warning_rows:
            row_index += 1
            row_index = write_section_header(row_index, "Warnings")
            row_index = write_table_header(
                row_index,
                ["Warning Type", "Details"],
            )
            for warning_row in request.projection.warning_rows:
                worksheet.cell(
                    row=row_index,
                    column=1,
                    value=warning_row.warning_type,
                )
                worksheet.cell(
                    row=row_index,
                    column=2,
                    value=warning_row.details,
                )
                for col_index in range(1, 3):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.border = table_border
                    cell.alignment = left_alignment
                row_index += 1

        row_index += 1
        row_index = write_section_header(row_index, "Results")
        row_index = write_table_header(
            row_index,
            ["Property", "Value", "Unit", "Notes"],
        )
        for result_row in request.projection.result_rows:
            worksheet.cell(
                row=row_index,
                column=1,
                value=result_row.property_name,
            )
            value = result_row.value
            value_cell = worksheet.cell(
                row=row_index,
                column=2,
                value=value if value != "" else None,
            )
            worksheet.cell(row=row_index, column=3, value=result_row.unit)
            worksheet.cell(row=row_index, column=4, value=result_row.notes)

            for col_index in range(1, 5):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment

            if isinstance(value, (int, float)):
                value_cell.alignment = right_alignment
                value_cell.number_format = get_excel_number_format(
                    result_row.unit
                )

            row_index += 1

        worksheet.column_dimensions["A"].width = 32
        worksheet.column_dimensions["B"].width = 22
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 52

        workbook.save(report_path)
        return report_path
