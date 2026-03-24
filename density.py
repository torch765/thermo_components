import sys
import sqlite3
import os # To check if DB file exists
from datetime import datetime

from chemicals import iapws as chemicals_iapws
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QComboBox, QListWidget,
    QPushButton, QLabel, QGridLayout, QVBoxLayout, QTableWidget,
    QTableWidgetItem, QMessageBox, QHeaderView, QHBoxLayout, QButtonGroup, QRadioButton, QStyleFactory
)
from PyQt6.QtCore import Qt, QTimer, QObject, pyqtSignal, QThread
from PyQt6.QtGui import QPalette, QColor

# --- Import the generated UI class ---
from gui import Ui_Dialog

# THERMO FLASH INTERFACE IMPORTS
# NOTE: Make sure PRMIX is imported if it's the only option for now
# If you add SRKMIX later, you'll need: from thermo.eos_mix import PRMIX, SRKMIX
from thermo import ChemicalConstantsPackage, CEOSGas, CEOSLiquid, FlashVL, PRMIX, FlashPureVLS, IAPWS95
# from thermo.eos_mix import PRMIX # Explicit import if needed


# --- Constants and Data Loading (Keep as is) ---
MOLECULAR_WEIGHTS = {
    "        ": 0.0,  # Dummy entry
    "hydrogen": 2.016, "carbon dioxide": 44.01, "carbon monoxide": 28.01,
    "hydrogen sulfide": 34.08, "methane": 16.04, "ethane": 30.07,
    "ethylene": 28.05, "propane": 44.10, "propylene": 42.08,
    "isobutane": 58.12, "n-butane": 58.12, "isobutylene": 56.11,
    "n-butylene": 56.11, "1-butene": 56.11, "2-butene": 56.11,
    "butadiene": 54.09, "isopentane": 72.15, "n-pentane": 72.15,
    "hexane": 86.18, "heptane": 100.20, "benzene": 78.11,
    "toluene": 92.14, "o-xylene": 106.17, "m-xylene": 106.17,
    "p-xylene": 106.17, "ethylbenzene": 106.17, "cumene": 120.19,
    "octane": 114.23, "nonane": 128.26, "oxygen": 32.00,
    "nitrogen": 28.01, "argon": 39.95, "water": 18.015,
}

R = 0.082057  # L·atm/(mol·K)
KCAL_PER_MJ = 238.845896627
MJ_PER_MMBTU = 1055.05585262
NORMAL_MOLAR_VOLUME_NM3_PER_KMOL = 22.414
NORMAL_T_C = 0.0
NORMAL_P_ATM = 1.0
STANDARD_T_F = 60.0
STANDARD_T_C = (STANDARD_T_F - 32.0) * 5.0 / 9.0
STANDARD_P_ATM = 1.0
ATM_TO_PA = 101325.0
KG_PER_TONNE = 1000.0
KG_PER_LB = 0.45359237
LB_PER_KLB = 1000.0
HOURS_PER_DAY = 24.0
M3_PER_BBL = 0.158987294928
FT3_PER_M3 = 35.3146667
M3_PER_FT3 = 0.028316846592

FLOW_UNIT_ORDER = [
    "kg/h", "kg/d", "t/h", "t/d", "lb/h", "lb/d", "Klb/h", "Klb/d",
    "Nm3/h", "Nm3/d", "Sm3/h", "Sm3/d", "bbl/h", "bbl/d", "SCFH", "MSCFD", "MMSCFD",
]

FLOW_UNIT_DEFINITIONS = {
    "kg/h": {"dimension": "mass", "basis": "none", "amount_to_base": 1.0, "time_to_day": HOURS_PER_DAY},
    "kg/d": {"dimension": "mass", "basis": "none", "amount_to_base": 1.0, "time_to_day": 1.0},
    "t/h": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_TONNE, "time_to_day": HOURS_PER_DAY},
    "t/d": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_TONNE, "time_to_day": 1.0},
    "lb/h": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_LB, "time_to_day": HOURS_PER_DAY},
    "lb/d": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_LB, "time_to_day": 1.0},
    "Klb/h": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_LB * LB_PER_KLB, "time_to_day": HOURS_PER_DAY},
    "Klb/d": {"dimension": "mass", "basis": "none", "amount_to_base": KG_PER_LB * LB_PER_KLB, "time_to_day": 1.0},
    "Nm3/h": {"dimension": "ref_volume", "basis": "normal", "amount_to_base": 1.0, "time_to_day": HOURS_PER_DAY},
    "Nm3/d": {"dimension": "ref_volume", "basis": "normal", "amount_to_base": 1.0, "time_to_day": 1.0},
    "Sm3/h": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": 1.0, "time_to_day": HOURS_PER_DAY},
    "Sm3/d": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": 1.0, "time_to_day": 1.0},
    "bbl/h": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": M3_PER_BBL, "time_to_day": HOURS_PER_DAY},
    "bbl/d": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": M3_PER_BBL, "time_to_day": 1.0},
    "SCFH": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": M3_PER_FT3, "time_to_day": HOURS_PER_DAY},
    "MSCFD": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": 1000.0 * M3_PER_FT3, "time_to_day": 1.0},
    "MMSCFD": {"dimension": "ref_volume", "basis": "standard", "amount_to_base": 1_000_000.0 * M3_PER_FT3, "time_to_day": 1.0},
}

WATER_COMPONENT_ALIASES = {"water", "h2o"}
PURE_WATER_WARNING_FRACTION = 0.999
PURE_WATER_ROUTE = "iapws95_pure_water"
PRMIX_DEFAULT_ROUTE = "prmix_default"
IAPWS95_MODEL_DISPLAY = "IAPWS-95"
IAPWS95_TWO_PHASE_REL_TOL = 1e-4
PRMIX_WATER_WARNING = (
    "Warning: Water is present. PRMIX results may be unreliable for aqueous or polar behavior. "
    "Use with caution."
)
PRMIX_TWO_PHASE_WATER_WARNING = (
    "Warning: Water-containing two-phase behavior may be unreliable with PRMIX. "
    "Validate density and phase behavior with a water-capable method."
)

def load_lhv_data(db_path='lhv_data.db'):
    """Loads LHV data from the SQLite database and returns it."""
    loaded_data = {}
    if not os.path.exists(db_path):
        print(f"Warning: LHV Database file not found at {db_path}. LHV calculations will not be available.")
        return loaded_data

    conn = None
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT component_name, lhv_mj_nm3 FROM lhv_values")
        rows = cursor.fetchall()
        for row in rows:
            loaded_data[row[0]] = row[1]
        print(f"Successfully loaded {len(loaded_data)} LHV entries from {db_path}.")
    except sqlite3.Error as e:
        print(f"Failed to load LHV data from database: {e}")
    finally:
        if conn:
            conn.close()
    return loaded_data

# Helper to find resource files in both dev and PyInstaller bundle
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def build_lhv_display_values(lhv_mj_nm3: float, mw_g_mol: float) -> dict:
    """Build display-ready LHV values from the base mixture MJ/Nm³ result."""
    try:
        lhv_mj_nm3 = float(lhv_mj_nm3)
    except (TypeError, ValueError):
        lhv_mj_nm3 = 0.0

    try:
        mw_g_mol = float(mw_g_mol)
    except (TypeError, ValueError):
        mw_g_mol = 0.0

    kcal_per_nm3 = lhv_mj_nm3 * KCAL_PER_MJ

    values = {
        "volumetric": {
            "MJ/Nm³": lhv_mj_nm3,
            "kcal/Nm³": kcal_per_nm3,
            "MMkcal/Nm³": kcal_per_nm3 / 1_000_000.0,
            "GJ/Nm³": lhv_mj_nm3 / 1000.0,
            "MMBtu/Nm³": lhv_mj_nm3 / MJ_PER_MMBTU,
        },
        "mass_basis": {
            "MJ/kg": None,
            "MJ/t": None,
            "GJ/kg": None,
            "GJ/t": None,
            "kcal/kg": None,
            "kcal/t": None,
            "MMkcal/kg": None,
            "MMkcal/t": None,
            "MMBtu/kg": None,
            "MMBtu/t": None,
        },
    }

    if mw_g_mol > 0:
        kg_per_nm3 = mw_g_mol / NORMAL_MOLAR_VOLUME_NM3_PER_KMOL
        if kg_per_nm3 > 0:
            mj_per_kg = lhv_mj_nm3 / kg_per_nm3
            mj_per_t = mj_per_kg * 1000.0
            kcal_per_kg = mj_per_kg * KCAL_PER_MJ
            kcal_per_t = kcal_per_kg * 1000.0
            values["mass_basis"] = {
                "MJ/kg": mj_per_kg,
                "MJ/t": mj_per_t,
                "GJ/kg": mj_per_kg / 1000.0,
                "GJ/t": mj_per_t / 1000.0,
                "kcal/kg": kcal_per_kg,
                "kcal/t": kcal_per_t,
                "MMkcal/kg": kcal_per_kg / 1_000_000.0,
                "MMkcal/t": kcal_per_t / 1_000_000.0,
                "MMBtu/kg": mj_per_kg / MJ_PER_MMBTU,
                "MMBtu/t": mj_per_t / MJ_PER_MMBTU,
            }

    # TODO: Add MW output once a flow input (for example Nm³/h or t/h) is introduced.
    return values


def format_lhv_display_value(value: float | None) -> str:
    """Format LHV values with readable precision for the results pane."""
    if value is None:
        return "N/A"
    decimals = 2 if abs(value) >= 1 else 4
    return f"{value:.{decimals}f}"


def get_excel_number_format(unit: str) -> str:
    """Return a sensible Excel number format for the given engineering unit."""
    if unit in {"Mol %", "Wt %"}:
        return "0.0000"
    if unit == "kg/m³":
        return "0.000"
    if unit in {"GJ/Nm³", "MMBtu/Nm³", "MMkcal/Nm³", "GJ/kg", "MMBtu/kg", "MMkcal/kg"}:
        return "0.0000"
    if unit:
        return "0.00"
    return "General"


def celsius_to_kelvin(temperature_c: float) -> float:
    return temperature_c + 273.15


def atm_to_pa(pressure_atm: float) -> float:
    return pressure_atm * ATM_TO_PA


def extract_scalar_density_value(density_result, phase, error) -> float | None:
    """Return a single density value only when the flash result is scalar."""
    if error or density_result is None:
        return None
    if phase == "Two-Phase" or isinstance(density_result, tuple):
        return None
    try:
        return float(density_result)
    except (TypeError, ValueError):
        return None


def build_density_note(phase, error) -> str:
    """Build a short note for density rows in the export report."""
    if error:
        return error
    if phase == "Two-Phase":
        return "Two-Phase result"
    if phase:
        return f"Phase: {phase}"
    return ""


def normalize_component_identity(name: str) -> str:
    """Normalize a component identifier for warning checks."""
    cleaned = str(name).strip().lower()
    compact = "".join(character for character in cleaned if character.isalnum())
    if compact == "h2o":
        return "water"
    return cleaned


def is_water_component(name: str) -> bool:
    """Return True when the component name maps to water."""
    return normalize_component_identity(name) in WATER_COMPONENT_ALIASES


def _active_basis_amount_rows(comp_names, mol_percents, wt_percents, basis: str) -> tuple[list[tuple[str, float]], float]:
    """Return positive active-basis amounts keyed by normalized component identity."""
    active_values = mol_percents if basis == "Mol %" else wt_percents
    rows = []
    total_active = 0.0

    for comp_name, raw_value in zip(comp_names, active_values):
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue

        if value <= 0:
            continue

        rows.append((normalize_component_identity(comp_name), value))
        total_active += value

    return rows, total_active


def has_water_component(comp_names, mol_percents, wt_percents, basis: str) -> bool:
    """Return True when water is present at a non-zero amount on the active basis."""
    rows, _ = _active_basis_amount_rows(comp_names, mol_percents, wt_percents, basis)
    return any(is_water_component(name) for name, _ in rows)


def water_fraction_active_basis(comp_names, mol_percents, wt_percents, basis: str) -> float:
    """Return the water fraction on the active input basis."""
    rows, total_active = _active_basis_amount_rows(comp_names, mol_percents, wt_percents, basis)
    if total_active <= 0:
        return 0.0
    water_active = sum(value for name, value in rows if is_water_component(name))
    return water_active / total_active


def is_effectively_pure_water(comp_names, mol_percents, wt_percents, basis: str) -> bool:
    """Return True when the active-basis composition is effectively pure water."""
    rows, total_active = _active_basis_amount_rows(comp_names, mol_percents, wt_percents, basis)
    if total_active <= 0:
        return False

    water_active = sum(value for name, value in rows if is_water_component(name))
    other_active = total_active - water_active
    if water_active <= 0:
        return False

    water_fraction = water_active / total_active
    other_fraction = other_active / total_active
    return water_fraction >= PURE_WATER_WARNING_FRACTION and other_fraction <= (1.0 - PURE_WATER_WARNING_FRACTION)


def select_thermo_route(comp_names, mol_percents, wt_percents, basis: str, default_eos: str) -> dict:
    """Select the thermo route for the current composition."""
    if is_effectively_pure_water(comp_names, mol_percents, wt_percents, basis):
        return {
            "route_id": PURE_WATER_ROUTE,
            "model_display": IAPWS95_MODEL_DISPLAY,
        }
    return {
        "route_id": PRMIX_DEFAULT_ROUTE,
        "model_display": str(default_eos or "PRMIX").strip() or "PRMIX",
    }


def phase_indicates_two_phase(phase) -> bool:
    """Return True when the phase text indicates two-phase behavior."""
    return "two-phase" in str(phase or "").strip().lower()


def build_thermo_warning_messages(comp_names, mol_percents, wt_percents, basis: str, thermo_route: str, result_data: dict) -> list[str]:
    """Build the persistent thermo warnings shown in the main tab and export."""
    if thermo_route != PRMIX_DEFAULT_ROUTE:
        return []

    if not has_water_component(comp_names, mol_percents, wt_percents, basis):
        return []

    messages = []
    messages.append(PRMIX_WATER_WARNING)

    has_two_phase_warning_condition = any(
        phase_indicates_two_phase(result_data.get(key))
        for key in ("phase", "density_normal_phase", "density_standard_phase")
    )
    if has_two_phase_warning_condition:
        messages.append(PRMIX_TWO_PHASE_WATER_WARNING)

    return messages


def parse_flow_input(text: str) -> float | None:
    """Parse a flow input, tolerating spaces and thousands separators."""
    cleaned = str(text).strip().replace(",", "").replace(" ", "")
    if not cleaned:
        return None
    return float(cleaned)


def format_flow_value(value: float) -> str:
    """Format flow results without scientific notation for typical engineering values."""
    formatted = f"{value:,.6f}"
    if "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def _coerce_positive_density(value) -> float | None:
    try:
        density = float(value)
    except (TypeError, ValueError):
        return None
    if density <= 0:
        return None
    return density


def _required_density_bases(from_meta: dict, to_meta: dict) -> set[str]:
    if from_meta["dimension"] == "mass" and to_meta["dimension"] == "mass":
        return set()
    if from_meta["dimension"] == "mass":
        return {to_meta["basis"]}
    if to_meta["dimension"] == "mass":
        return {from_meta["basis"]}
    if from_meta["basis"] == to_meta["basis"]:
        return set()
    return {from_meta["basis"], to_meta["basis"]}


def _missing_density_message(missing_bases: set[str]) -> str:
    if missing_bases == {"normal"}:
        return "Normal density required"
    if missing_bases == {"standard"}:
        return "Standard density required"
    return "Normal and standard densities required"


def convert_flow(value, from_unit, to_unit, density_normal_kg_m3=None, density_standard_kg_m3=None):
    """Convert between mass and reference-volume flow units using density bridges."""
    from_meta = FLOW_UNIT_DEFINITIONS.get(from_unit)
    to_meta = FLOW_UNIT_DEFINITIONS.get(to_unit)
    if from_meta is None or to_meta is None:
        raise ValueError("Select source and destination units")

    value = float(value)
    required_bases = _required_density_bases(from_meta, to_meta)
    density_map = {
        "normal": _coerce_positive_density(density_normal_kg_m3),
        "standard": _coerce_positive_density(density_standard_kg_m3),
    }
    missing_bases = {basis for basis in required_bases if density_map.get(basis) is None}
    if missing_bases:
        raise ValueError(_missing_density_message(missing_bases))

    from_base_per_day = value * from_meta["amount_to_base"] * from_meta["time_to_day"]

    if from_meta["dimension"] == "mass":
        mass_kg_day = from_base_per_day
        if to_meta["dimension"] == "mass":
            return mass_kg_day / (to_meta["amount_to_base"] * to_meta["time_to_day"])
        target_volume_m3_day = mass_kg_day / density_map[to_meta["basis"]]
        return target_volume_m3_day / (to_meta["amount_to_base"] * to_meta["time_to_day"])

    source_volume_m3_day = from_base_per_day
    if to_meta["dimension"] == "ref_volume" and from_meta["basis"] == to_meta["basis"]:
        return source_volume_m3_day / (to_meta["amount_to_base"] * to_meta["time_to_day"])

    mass_kg_day = source_volume_m3_day * density_map[from_meta["basis"]]
    if to_meta["dimension"] == "mass":
        return mass_kg_day / (to_meta["amount_to_base"] * to_meta["time_to_day"])

    target_volume_m3_day = mass_kg_day / density_map[to_meta["basis"]]
    return target_volume_m3_day / (to_meta["amount_to_base"] * to_meta["time_to_day"])

# Step 1: Worker class for threaded calculation
class CalculationWorker(QObject):
    result = pyqtSignal(dict)
    error = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, calculator, lhv_data, comp_names, mol_percents, wt_percents, basis, T_k, pressure_pa, pressure_atm):
        super().__init__()
        self.calculator = calculator
        self.lhv_data = lhv_data
        self.comp_names = comp_names
        self.mol_percents = mol_percents
        self.wt_percents = wt_percents
        self.basis = basis
        self.T_k = T_k
        self.pressure_pa = pressure_pa
        self.pressure_atm = pressure_atm

    def run(self):
        try:
            # --- Convert to mole fractions based on input basis ---
            mole_fracs_dict = {}
            if self.basis == "Mol %":
                total_mol_percent = sum(self.mol_percents)
                if abs(total_mol_percent - 100.0) > 1e-4 or total_mol_percent == 0:
                    self.error.emit("Error: Invalid Mol % input.")
                    self.finished.emit()
                    return
                for name, percent in zip(self.comp_names, self.mol_percents):
                    mole_fracs_dict[name] = percent / 100.0
            else:
                total_wt_percent = sum(self.wt_percents)
                if abs(total_wt_percent - 100.0) > 1e-4 or total_wt_percent == 0:
                    self.error.emit("Error: Invalid Wt % input.")
                    self.finished.emit()
                    return
                moles = {}
                total_moles = 0.0
                for name, percent in zip(self.comp_names, self.wt_percents):
                    mw = MOLECULAR_WEIGHTS.get(name)
                    if mw is None or mw <= 0:
                        self.error.emit(f"Error: Missing or invalid MW for {name}.")
                        self.finished.emit()
                        return
                    moles[name] = (percent / 100.0) * 100.0 / mw
                    total_moles += moles[name]
                if total_moles == 0:
                    self.error.emit("Error: Total moles is zero after Wt% conversion.")
                    self.finished.emit()
                    return
                for name in self.comp_names:
                    mole_fracs_dict[name] = moles[name] / total_moles

            self.calculator.set_components(mole_fracs_dict)
            thermo_route = select_thermo_route(
                self.comp_names,
                self.mol_percents,
                self.wt_percents,
                self.basis,
                self.calculator.eos,
            )

            # --- Perform Calculations ---
            mw = self.calculator.calculate_molecular_weight()
            density_result, phase, error = self.calculator.calculate_density_for_route(
                self.T_k,
                self.pressure_pa,
                thermo_route["route_id"],
            )
            density_normal_result, density_normal_phase, density_normal_error = self.calculator.calculate_density_for_route(
                celsius_to_kelvin(NORMAL_T_C),
                atm_to_pa(NORMAL_P_ATM),
                thermo_route["route_id"],
            )
            density_standard_result, density_standard_phase, density_standard_error = self.calculator.calculate_density_for_route(
                celsius_to_kelvin(STANDARD_T_C),
                atm_to_pa(STANDARD_P_ATM),
                thermo_route["route_id"],
            )
            bubble_point, bp_error = self.calculator.calculate_bubble_point_for_route(
                self.pressure_pa,
                thermo_route["route_id"],
            )
            mixture_lhv, missing_lhv = self.calculator.calculate_lhv(self.lhv_data)

            result_data = {
                'mw': mw,
                'comp_names': list(self.comp_names),
                'mol_percents': list(self.mol_percents),
                'wt_percents': list(self.wt_percents),
                'thermo_route': thermo_route["route_id"],
                'model_display': thermo_route["model_display"],
                'density_result': density_result,
                'phase': phase,
                'density_error': error,
                'density_actual_kg_m3': extract_scalar_density_value(density_result, phase, error),
                'density_normal_result': density_normal_result,
                'density_normal_phase': density_normal_phase,
                'density_normal_error': density_normal_error,
                'density_normal_kg_m3': extract_scalar_density_value(
                    density_normal_result, density_normal_phase, density_normal_error
                ),
                'density_standard_result': density_standard_result,
                'density_standard_phase': density_standard_phase,
                'density_standard_error': density_standard_error,
                'density_standard_kg_m3': extract_scalar_density_value(
                    density_standard_result, density_standard_phase, density_standard_error
                ),
                'bubble_point': bubble_point,
                'bp_error': bp_error,
                'mixture_lhv': mixture_lhv,
                'missing_lhv': missing_lhv,
                'basis': self.basis,
                'eos': self.calculator.eos,
                'pressure_atm': self.pressure_atm,
            }
            result_data['warnings'] = build_thermo_warning_messages(
                self.comp_names,
                self.mol_percents,
                self.wt_percents,
                self.basis,
                thermo_route["route_id"],
                result_data,
            )
            self.result.emit(result_data)
        except Exception as e:
            import traceback
            self.error.emit(f"Worker Exception: {e}\n{traceback.format_exc()}")
        self.finished.emit()

# --- MixtureCalculator Class (do not change this code) ---
class MixtureCalculator:
    """Handles the thermodynamic calculations, decoupled from the UI."""
    def __init__(self, components=None, eos='PRMIX'):
        self.components = components if components is not None else {}
        self.eos = eos
        self.constants = None
        self.properties = None

    def set_components(self, components):
        self.components = components
        self.constants = None # Reset cached constants when components change
        self.properties = None # Reset cached properties

    def set_eos(self, eos):
        # Add logic here if you support more EOS types later
        self.eos = eos
        print(f"EOS set to: {self.eos}") # Debug print

    def calculate_molecular_weight(self):
        if not self.components: return 0.0
        mw_sum = 0.0
        frac_sum = sum(self.components.values())
        if frac_sum == 0: return 0.0

        normalized_components = {comp: frac / frac_sum for comp, frac in self.components.items()}

        for comp_name, mole_frac in normalized_components.items():
            mw = MOLECULAR_WEIGHTS.get(comp_name, 0.0)
            mw_sum += mw * mole_frac
        # The average MW doesn't depend on the sum of fractions if normalized
        return mw_sum # / frac_sum if frac_sum > 0 else 0.0 -> removed as normalized

    def calculate_density_for_route(self, temperature_k, pressure_pa, route_id):
        """Dispatch density calculation to the selected thermo route."""
        if route_id == PURE_WATER_ROUTE:
            return self.calculate_pure_water_density_iapws95(temperature_k, pressure_pa)
        return self.calculate_density(temperature_k, pressure_pa)

    def calculate_bubble_point_for_route(self, pressure_pa, route_id):
        """Dispatch bubble-point/saturation calculation to the selected thermo route."""
        if route_id == PURE_WATER_ROUTE:
            return self.calculate_pure_water_bubble_point_iapws95(pressure_pa)
        return self.calculate_bubble_point(pressure_pa)

    def calculate_pure_water_density_iapws95(self, temperature_k, pressure_pa):
        """Calculate pure-water density using the local IAPWS-95 implementation."""
        if temperature_k is None or pressure_pa is None:
            return None, None, "Missing temperature or pressure."
        if temperature_k <= 0 or pressure_pa <= 0:
            return None, None, "Temperature and pressure must be positive."

        try:
            phase_label = None
            critical_t = IAPWS95.Tc
            critical_p = IAPWS95.Pc

            if 235.0 <= temperature_k < critical_t and pressure_pa < critical_p:
                psat = chemicals_iapws.iapws95_Psat(temperature_k)
                relative_difference = abs(pressure_pa - psat) / max(psat, 1.0)
                if relative_difference <= IAPWS95_TWO_PHASE_REL_TOL:
                    density_liq = chemicals_iapws.iapws95_rhol_sat(temperature_k)
                    density_gas = chemicals_iapws.iapws95_rhog_sat(temperature_k)
                    return (density_liq, density_gas), "Two-Phase", None
                phase_label = "Liquid" if pressure_pa > psat else "Vapor"
            elif temperature_k >= critical_t and pressure_pa >= critical_p:
                phase_label = "Supercritical"
            elif temperature_k >= critical_t:
                phase_label = "Vapor"
            elif pressure_pa >= critical_p:
                phase_label = "Liquid"

            water_state = IAPWS95(T=temperature_k, P=pressure_pa, zs=[1.0])
            density_mass = water_state.rho_mass()

            if phase_label is None:
                phase_label = "Liquid" if density_mass >= 200.0 else "Vapor"

            return density_mass, phase_label, None
        except Exception as exc:
            return None, None, f"IAPWS-95 pure-water density failed: {exc}"

    def calculate_pure_water_bubble_point_iapws95(self, pressure_pa):
        """Calculate pure-water saturation temperature using IAPWS-95."""
        if pressure_pa is None or pressure_pa <= 0:
            return None, "Pressure must be positive."

        try:
            saturation_temperature_k = chemicals_iapws.iapws95_Tsat(pressure_pa)
        except Exception as exc:
            return None, f"IAPWS-95 saturation calculation failed: {exc}"

        return saturation_temperature_k - 273.15, None


    def calculate_density(self, temperature_k, pressure_pa):
        """Calculates density, accounting for phase (liquid/vapor/two-phase).
           (Reverted to original phase iteration logic)"""
        if not self.components:
            # Return format matches original expectation: density, phase_str, error_str
            return None, None, "No components selected."

        total = sum(self.components.values())
        if total == 0:
            return None, None, "Total fraction is zero."

        zs = [x / total for x in self.components.values()]
        comp_list = list(self.components.keys())
        print(f"Calculating density (original logic) for: {comp_list} at T={temperature_k:.2f} K, P={pressure_pa:.1f} Pa with EOS={self.eos}")

        try:
            if self.constants is None or self.properties is None:
                print("Fetching chemical constants...")
                self.constants, self.properties = ChemicalConstantsPackage.from_IDs(comp_list)
                print("Constants fetched.")

            # Basic check for constants needed by EOS
            if None in self.constants.Tcs or None in self.constants.Pcs or None in self.constants.omegas:
                 missing_data_comps = [comp_list[i] for i, tc in enumerate(self.constants.Tcs) if tc is None]
                 return None, None, f"Missing critical properties for: {', '.join(missing_data_comps)}"

            eos_kwargs = {
                'Pcs': self.constants.Pcs,
                'Tcs': self.constants.Tcs,
                'omegas': self.constants.omegas,
            }

            eos_class = PRMIX # Assuming PRMIX only for now
            hcap_gases = self.properties.HeatCapacityGases if hasattr(self.properties, 'HeatCapacityGases') else None

            if len(comp_list) == 1:
                 # Original code used FlashPureVLS here, keep that part
                 gas = CEOSGas(eos_class, eos_kwargs=eos_kwargs, HeatCapacityGases=hcap_gases)
                 liquid = CEOSLiquid(eos_class, eos_kwargs=eos_kwargs, HeatCapacityGases=hcap_gases)
                 flasher = FlashPureVLS(self.constants, self.properties, gas=gas, liquids=[liquid], solids=[])
            else:
                 gas = CEOSGas(eos_class, HeatCapacityGases=hcap_gases, eos_kwargs=eos_kwargs, zs=zs)
                 liquid = CEOSLiquid(eos_class, HeatCapacityGases=hcap_gases, eos_kwargs=eos_kwargs, zs=zs)
                 flasher = FlashVL(self.constants, self.properties, liquid=liquid, gas=gas)

            print(f"Performing flash T={temperature_k}, P={pressure_pa}, zs={zs}")
            result = flasher.flash(T=temperature_k, P=pressure_pa, zs=zs)
            # Debug: Print result attributes relevant to original logic
            print(f"Flash result: Phases present = {len(result.phases)}, PurePhase='{getattr(result, 'phase', 'N/A')}'")
            if hasattr(result, 'phases'):
                 for i, p in enumerate(result.phases):
                      print(f"  Phase {i}: Liquid={p.is_liquid}, Gas={p.is_gas}, V={p.V()}")

            # Calculate Overall Molecular Weight (convert g/mol to kg/mol) - used by original logic
            MWs = self.constants.MWs if self.constants is not None else []
            if not MWs or None in MWs:
                return None, None, "Could not retrieve molecular weight for all components."
             # Use overall mole fractions (zs) for average MW calculation
            MW = sum([zi * Mwi for zi, Mwi in zip(zs, MWs)]) / 1000.0
            if MW <= 0:
                 return None, None, "Calculated average molecular weight is zero or negative."


            density_liq = None
            density_gas = None

            # --- Original Logic based on iterating result.phases (for mixtures) ---
            # --- and result.phase (for pure) ---
            if len(comp_list) == 1 and isinstance(flasher, FlashPureVLS):
                 # Handle pure component flash results based on 'phase' attribute
                 pure_phase_attr = getattr(result, 'phase', '').lower()
                 if pure_phase_attr in ['l', 'liquid']:
                      if result.liquid0 and result.liquid0.V() is not None and result.liquid0.V() > 0:
                           density_liq = MW / result.liquid0.V()
                 elif pure_phase_attr in ['g', 'v', 'gas', 'vapor']:
                      if result.gas and result.gas.V() is not None and result.gas.V() > 0:
                           density_gas = MW / result.gas.V()
                 elif '/' in pure_phase_attr: # Two phase like 'l/g'
                      if result.liquid0 and result.liquid0.V() is not None and result.liquid0.V() > 0:
                           density_liq = MW / result.liquid0.V()
                      if result.gas and result.gas.V() is not None and result.gas.V() > 0:
                           density_gas = MW / result.gas.V()
                 else:
                      # If phase attribute is unexpected, try checking phases list
                      if hasattr(result, 'phases'):
                           for phase in result.phases:
                                if phase.is_liquid and phase.V() is not None and phase.V() > 0:
                                    density_liq = MW / phase.V()
                                elif phase.is_gas and phase.V() is not None and phase.V() > 0:
                                    density_gas = MW / phase.V()
                      else:
                          return None, None, f"Unexpected pure phase result: {pure_phase_attr}"


            elif hasattr(result, 'phases'): # Mixture case: Iterate phases list
                 for phase in result.phases:
                      phase_vol = phase.V()
                      if phase_vol is not None and phase_vol > 0:
                           # Original logic used overall MW for phase density estimate
                           if phase.is_liquid:
                                density_liq = MW / phase_vol
                           elif phase.is_gas:
                                density_gas = MW / phase_vol
                      else:
                           print(f"Warning: Phase volume is None or zero for phase type (Liq={phase.is_liquid}, Gas={phase.is_gas})")

            else:
                 return None, None, "Flash result object did not contain expected 'phases' attribute."

            # --- Determine final return based on calculated densities ---
            if density_liq is not None and density_gas is not None:
                print(f"Determined Phase: Two-Phase, DensLiq={density_liq:.3f}, DensGas={density_gas:.3f}")
                return (density_liq, density_gas), "Two-Phase", None
            elif density_liq is not None:
                print(f"Determined Phase: Liquid, DensLiq={density_liq:.3f}")
                return density_liq, "Liquid", None
            elif density_gas is not None:
                print(f"Determined Phase: Vapor, DensGas={density_gas:.3f}")
                return density_gas, "Vapor", None
            else:
                # If no densities could be calculated
                print(f"Warning: Could not calculate density for any phase found.")
                # Try to determine phase based on VF if available, otherwise return Unknown
                phase_guess = "Unknown"
                if hasattr(result, 'VF'):
                     if result.VF == 0: phase_guess = "Liquid (calc failed)"
                     elif result.VF == 1: phase_guess = "Vapor (calc failed)"
                     elif 0 < result.VF < 1: phase_guess = "Two-Phase (calc failed)"
                return None, phase_guess, "Failed to calculate density values."

        except Exception as e:
            import traceback
            print(f"Error during density calculation: {e}\n{traceback.format_exc()}")
            # Return format matches original expectation
            return None, None, str(e)
    
    def calculate_bubble_point(self, pressure_pa):
        """Calculates the bubble point temperature."""
        if not self.components: return None, "No components selected."
        total = sum(self.components.values())
        if total == 0: return None, "Total fraction is zero."

        zs = [x / total for x in self.components.values()]
        comp_list = list(self.components.keys())
        print(f"Calculating bubble point for: {comp_list} at P={pressure_pa:.1f} Pa")

        try:
            if self.constants is None or self.properties is None:
                 self.constants, self.properties = ChemicalConstantsPackage.from_IDs(comp_list)

            if None in self.constants.Tcs or None in self.constants.Pcs or None in self.constants.omegas:
                 missing_data_comps = [comp_list[i] for i, tc in enumerate(self.constants.Tcs) if tc is None]
                 return None, f"Missing critical properties for: {', '.join(missing_data_comps)}"

            eos_kwargs = {'Pcs': self.constants.Pcs, 'Tcs': self.constants.Tcs, 'omegas': self.constants.omegas,}
            hcap_gases = self.properties.HeatCapacityGases if hasattr(self.properties, 'HeatCapacityGases') else None
            eos_class = PRMIX # Assuming PRMIX for now

            if len(comp_list) == 1:
                gas = CEOSGas(eos_class, eos_kwargs=eos_kwargs, HeatCapacityGases=hcap_gases)
                liquid = CEOSLiquid(eos_class, eos_kwargs=eos_kwargs, HeatCapacityGases=hcap_gases)
                flasher = FlashPureVLS(self.constants, self.properties, gas=gas, liquids=[liquid], solids=[])
            else:
                gas = CEOSGas(eos_class, HeatCapacityGases=hcap_gases, eos_kwargs=eos_kwargs, zs=zs)
                liquid = CEOSLiquid(eos_class, HeatCapacityGases=hcap_gases, eos_kwargs=eos_kwargs, zs=zs)
                flasher = FlashVL(self.constants, self.properties, liquid=liquid, gas=gas)

            result = flasher.flash(P=pressure_pa, VF=0, zs=zs) # VF=0 for bubble point
            print(f"Bubble point flash result T={result.T}")
            return result.T - 273.15, None

        except Exception as e:
             import traceback
             print(f"Error during bubble point calculation: {e}\n{traceback.format_exc()}")
             return None, str(e)

    def calculate_lhv(self, lhv_data):
        """Calculates the LHV of the mixture based on mole fractions (MJ/Nm³)."""
        if not self.components: return 0.0, []
        mixture_lhv = 0.0
        missing_components = []
        total_fraction = sum(self.components.values())
        if total_fraction == 0: return 0.0, []

        normalized_components = {comp: frac / total_fraction for comp, frac in self.components.items()}
        for comp_name, mole_frac in normalized_components.items():
            component_lhv = lhv_data.get(comp_name)
            if component_lhv is not None:
                mixture_lhv += mole_frac * component_lhv
            else:
                if comp_name.strip(): missing_components.append(comp_name)
        return mixture_lhv, missing_components


# --- MainWindow Class (Modified for gui.py) ---
class MainWindow(QMainWindow):
    def __init__(self, lhv_data):
        super().__init__()

        # --- Set up the UI from Designer ---
        self.ui = Ui_Dialog()
        self.ui.setupUi(self) # Setup the UI onto this QMainWindow
        if hasattr(self.ui, "tabWidget"):
            self.ui.tabWidget.setCurrentIndex(0)

        # Step 2: Progress bar animation state
        self.progress_timer = None
        self.progress_target = 100
        self.progress_animation_duration = 800  # milliseconds
        self.progress_animation_start_time = 0
        self.progress_animation_start_value = 0

        # --- Store LHV data and calculator instance ---
        self.lhv_database = lhv_data
        self.lhv_data_loaded = bool(self.lhv_database)
        self.calculator = MixtureCalculator()
        self.last_result_data = None
        self.density_actual_kg_m3 = None
        self.density_normal_kg_m3 = None
        self.density_standard_kg_m3 = None

        if not self.lhv_data_loaded:
             print("Warning: MainWindow initialized with no LHV data.")
             # You could disable LHV display or show a warning label here if needed

        # --- Populate widgets ---
        self.populate_comboboxes()
        self.setup_table()
        self.setup_thermo_warning_banner()
        self.setup_flow_tab()

        # --- Connect signals ---
        self.connect_signals()

        # --- Initialize UI State ---
        self.ui.radioButton_mol_percent.setChecked(True) # Default Mol%
        self.on_input_basis_changed() # Set initial table state

        # "Export Report" would be more accurate, but keep "Print Results" as requested.
        if hasattr(self.ui, "printResultsButton"):
            self.ui.printResultsButton.setEnabled(True)

        # Set Window Title (Optional - can also be set in Designer)
        self.setWindowTitle("Thermo Calculator - v.1.0")

    def _parse_float_or_zero(self, text: str) -> float:
        """Parse float from a table cell; treat blanks/invalid as 0.0."""
        try:
            return float(text) if text and str(text).strip() else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _recalculate_inactive_column(self):
        """Auto-calculate the inactive composition column from the active one.

        - If Mol% is active, compute Wt% (inactive).
        - If Wt% is active, compute Mol% (inactive).
        The inactive column remains read-only and gray.
        """
        row_count = self.ui.tableWidget.rowCount()
        if row_count <= 1:
            return

        is_mol_basis = self.ui.radioButton_mol_percent.isChecked()
        active_col = 1 if is_mol_basis else 2
        inactive_col = 2 if is_mol_basis else 1
        total_row = row_count - 1

        names = []
        active_vals = []
        mws = []

        for row in range(total_row):
            name_item = self.ui.tableWidget.item(row, 0)
            if not name_item:
                continue
            name = name_item.text()
            names.append(name)

            active_item = self.ui.tableWidget.item(row, active_col)
            active_vals.append(self._parse_float_or_zero(active_item.text() if active_item else ""))

            mw = MOLECULAR_WEIGHTS.get(name, 0.0)
            mws.append(mw if mw is not None else 0.0)

        derived = [None] * len(names)
        if is_mol_basis:
            # Mol% -> Wt%
            masses = []
            total_mass = 0.0
            for mol, mw in zip(active_vals, mws):
                m = (mol * mw) if (mw and mw > 0 and mol != 0) else 0.0
                masses.append(m)
                total_mass += m
            if total_mass > 0:
                derived = [(100.0 * m / total_mass) for m in masses]
        else:
            # Wt% -> Mol%
            moles = []
            total_moles = 0.0
            for wt, mw in zip(active_vals, mws):
                n = (wt / mw) if (mw and mw > 0 and wt != 0) else 0.0
                moles.append(n)
                total_moles += n
            if total_moles > 0:
                derived = [(100.0 * n / total_moles) for n in moles]

        # Write derived values into inactive column (avoid recursion)
        self.ui.tableWidget.blockSignals(True)
        try:
            name_idx = 0
            for row in range(total_row):
                name_item = self.ui.tableWidget.item(row, 0)
                if not name_item:
                    continue

                inactive_item = self.ui.tableWidget.item(row, inactive_col)
                if inactive_item is None:
                    inactive_item = QTableWidgetItem("")
                    self.ui.tableWidget.setItem(row, inactive_col, inactive_item)

                val = derived[name_idx]
                inactive_item.setText("" if val is None else f"{val:.4f}")

                # Keep inactive styling/flags consistent
                inactive_item.setBackground(QColor("lightGray"))
                inactive_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)

                name_idx += 1

            # Keep inactive total cell blank
            total_inactive_item = self.ui.tableWidget.item(total_row, inactive_col)
            if total_inactive_item is not None:
                total_inactive_item.setText("")
                total_inactive_item.setBackground(QColor("lightGray"))
                total_inactive_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
        finally:
            self.ui.tableWidget.blockSignals(False)

    def invalidate_results(self, clear_visible_results: bool = True):
        """Clear the latest successful result so exports cannot use stale data."""
        self.last_result_data = None
        self.density_actual_kg_m3 = None
        self.density_normal_kg_m3 = None
        self.density_standard_kg_m3 = None
        self.set_thermo_warning_messages([])
        if clear_visible_results:
            self.ui.results_list.clear()
        self.update_flow_conversion()

    def setup_thermo_warning_banner(self):
        """Create a persistent, non-modal warning banner above the main results area."""
        self.thermo_warning_label = QLabel(self.ui.tab)
        self.thermo_warning_label.setObjectName("thermoWarningLabel")
        self.thermo_warning_label.setWordWrap(True)
        self.thermo_warning_label.setTextFormat(Qt.TextFormat.PlainText)
        self.thermo_warning_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        self.thermo_warning_label.setMargin(8)
        self.thermo_warning_label.setStyleSheet(
            "QLabel#thermoWarningLabel {"
            "background-color: rgb(255, 244, 204);"
            "color: rgb(122, 63, 0);"
            "border: 1px solid rgb(230, 184, 0);"
            "border-radius: 4px;"
            "}"
        )
        self.thermo_warning_label.hide()

        self._results_label_base_geometry = self.ui.results_label.geometry()
        self._results_list_base_geometry = self.ui.results_list.geometry()
        self._progress_bar_base_geometry = self.ui.progressBar.geometry() if hasattr(self.ui, "progressBar") else None
        self._warning_banner_y = self.ui.groupBox.geometry().bottom() + 8

    def set_thermo_warning_messages(self, warning_messages: list[str]):
        """Show or hide the persistent thermo warning banner."""
        if not hasattr(self, "thermo_warning_label"):
            return

        cleaned_messages = [str(message).strip() for message in warning_messages if str(message).strip()]
        if not cleaned_messages:
            self.thermo_warning_label.hide()
            self.ui.results_label.setGeometry(self._results_label_base_geometry)
            self.ui.results_list.setGeometry(self._results_list_base_geometry)
            if self._progress_bar_base_geometry is not None:
                self.ui.progressBar.setGeometry(self._progress_bar_base_geometry)
            return

        warning_text = "\n".join(cleaned_messages)
        banner_width = self._results_list_base_geometry.width()
        text_rect = self.thermo_warning_label.fontMetrics().boundingRect(
            0,
            0,
            banner_width - 16,
            1000,
            int(Qt.TextFlag.TextWordWrap),
            warning_text,
        )
        banner_height = max(44, text_rect.height() + 16)
        self.thermo_warning_label.setText(warning_text)
        self.thermo_warning_label.setGeometry(
            self._results_list_base_geometry.x(),
            self._warning_banner_y,
            banner_width,
            banner_height,
        )
        self.thermo_warning_label.show()
        self.thermo_warning_label.raise_()

        results_label_y = self._warning_banner_y + banner_height + 8
        self.ui.results_label.setGeometry(
            self._results_label_base_geometry.x(),
            results_label_y,
            self._results_label_base_geometry.width(),
            self._results_label_base_geometry.height(),
        )

        list_y_offset = self._results_list_base_geometry.y() - self._results_label_base_geometry.y()
        results_list_y = results_label_y + list_y_offset
        base_results_bottom = self._results_list_base_geometry.y() + self._results_list_base_geometry.height()
        results_list_height = max(100, base_results_bottom - results_list_y)
        self.ui.results_list.setGeometry(
            self._results_list_base_geometry.x(),
            results_list_y,
            self._results_list_base_geometry.width(),
            results_list_height,
        )

        if self._progress_bar_base_geometry is not None:
            self.ui.progressBar.setGeometry(
                self._progress_bar_base_geometry.x(),
                base_results_bottom + 8,
                self._progress_bar_base_geometry.width(),
                self._progress_bar_base_geometry.height(),
            )

    def get_export_base_dir(self):
        """Return the directory where generated reports should be saved."""
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))

    def build_report_filename(self, timestamp: datetime | None = None):
        """Build the timestamped Excel report filename."""
        timestamp = timestamp or datetime.now()
        return f"Thermo_Report_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"

    def get_current_conditions(self):
        """Return the currently active calculation conditions for reporting."""
        basis = self.last_result_data.get("basis") if self.last_result_data else (
            "Mol %" if self.ui.radioButton_mol_percent.isChecked() else "Wt %"
        )
        model_display = self.last_result_data.get("model_display") if self.last_result_data else (
            self.ui.comboBox_select_EOS.currentText()
        )
        return [
            ("Basis", basis),
            ("Temperature", self.ui.comboBox_select_temperature.currentText()),
            ("Pressure", self.ui.comboBox_select_pressure.currentText()),
            ("Model / EOS", model_display),
        ]

    def _coerce_report_value(self, text: str):
        """Convert report cell text to float when possible, preserving blanks/text."""
        cleaned = str(text).strip() if text is not None else ""
        if not cleaned:
            return ""
        try:
            return float(cleaned)
        except ValueError:
            return cleaned

    def get_input_composition_rows(self):
        """Return the input composition table rows, including the Total row."""
        rows = []
        for row_index in range(self.ui.tableWidget.rowCount()):
            component_item = self.ui.tableWidget.item(row_index, 0)
            mol_item = self.ui.tableWidget.item(row_index, 1)
            wt_item = self.ui.tableWidget.item(row_index, 2)
            rows.append({
                "Component": component_item.text() if component_item else "",
                "Mol %": self._coerce_report_value(mol_item.text() if mol_item else ""),
                "Wt %": self._coerce_report_value(wt_item.text() if wt_item else ""),
            })
        return rows

    def build_report_warning_rows(self, result_data):
        """Build structured warning rows for the export report."""
        warning_rows = []

        for warning_message in result_data.get("warnings") or []:
            warning_rows.append({
                "Warning Type": "Thermo",
                "Details": warning_message,
            })

        missing_lhv = result_data.get("missing_lhv") or []
        if missing_lhv:
            warning_rows.append({
                "Warning Type": "LHV",
                "Details": f"LHV Warning: No data for: {', '.join(missing_lhv)}",
            })

        return warning_rows

    def build_results_rows(self, result_data):
        """Build structured result rows for report export without scraping the UI."""
        rows = []

        def add_row(property_name, value, unit="", notes=""):
            rows.append({
                "Property": property_name,
                "Value": value,
                "Unit": unit,
                "Notes": notes,
            })

        add_row("Average molecular weight", result_data.get("mw"), "g/mol")

        phase = result_data.get("phase")
        density_result = result_data.get("density_result")
        density_error = result_data.get("density_error")
        if phase:
            add_row("Phase @ selected conditions", phase)
        else:
            add_row("Phase @ selected conditions", "Could not determine", notes=density_error or "")

        if density_error:
            add_row("Density @ selected conditions", "N.A.", "kg/m³", density_error)
        elif phase == "Two-Phase" and isinstance(density_result, tuple) and len(density_result) == 2:
            density_liq, density_gas = density_result
            add_row(
                "Density @ selected conditions (Liq)",
                density_liq if density_liq is not None else "N.A.",
                "kg/m³",
            )
            add_row(
                "Density @ selected conditions (Vap)",
                density_gas if density_gas is not None else "N.A.",
                "kg/m³",
            )
        elif density_result is not None:
            add_row("Density @ selected conditions", density_result, "kg/m³")
        else:
            add_row("Density @ selected conditions", "N.A.", "kg/m³")

        def add_reference_density_row(property_name, density_value, phase_value, error_value):
            add_row(
                property_name,
                density_value if density_value is not None else "N.A.",
                "kg/m³",
                build_density_note(phase_value, error_value),
            )

        add_reference_density_row(
            "Density @ normal conditions",
            result_data.get("density_normal_kg_m3"),
            result_data.get("density_normal_phase"),
            result_data.get("density_normal_error"),
        )
        add_reference_density_row(
            "Density @ standard conditions",
            result_data.get("density_standard_kg_m3"),
            result_data.get("density_standard_phase"),
            result_data.get("density_standard_error"),
        )

        pressure_atm = result_data.get("pressure_atm")
        bubble_point_label = f"Bubble point @ {pressure_atm:g} atm" if pressure_atm is not None else "Bubble point"
        bubble_point = result_data.get("bubble_point")
        bp_error = result_data.get("bp_error")
        bubble_point_note = "IAPWS-95 saturation temperature for pure water." if result_data.get("thermo_route") == PURE_WATER_ROUTE else ""
        if bp_error:
            add_row(bubble_point_label, "N/A", "°C", bp_error)
        elif bubble_point is not None:
            add_row(bubble_point_label, bubble_point, "°C", bubble_point_note)
        else:
            add_row(bubble_point_label, "Calculation Failed", "°C", bubble_point_note)

        if self.lhv_data_loaded:
            lhv_display_values = build_lhv_display_values(result_data.get("mixture_lhv", 0.0), result_data.get("mw", 0.0))
            add_row("Mixture LHV", lhv_display_values["volumetric"]["MJ/Nm³"], "MJ/Nm³", "Base value")
            for unit in ["kcal/Nm³", "MMkcal/Nm³", "GJ/Nm³", "MMBtu/Nm³"]:
                add_row("Mixture LHV", lhv_display_values["volumetric"][unit], unit)

            mass_note = "Derived from mixture MW and 22.414 Nm³/kmol at 0 °C and 1 atm."
            for index, unit in enumerate([
                "MJ/kg", "MJ/t", "GJ/kg", "GJ/t", "kcal/kg", "kcal/t",
                "MMkcal/kg", "MMkcal/t", "MMBtu/kg", "MMBtu/t"
            ]):
                add_row(
                    "Mixture LHV (Mass basis)",
                    lhv_display_values["mass_basis"][unit],
                    unit,
                    mass_note if index == 0 else "",
                )
        else:
            add_row("LHV data", "N/A", "", "LHV database not loaded.")

        return rows

    def export_results_to_excel(self):
        """Export the latest calculation results to a formatted Excel workbook."""
        if not self.last_result_data:
            QMessageBox.information(self, "Print Results", "No calculation results available. Please click Go first.")
            return None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            QMessageBox.warning(
                self,
                "Print Results",
                "Excel export requires the 'openpyxl' package.\n"
                "Install it in this environment and try again.\n\n"
                "Command:\npython -m pip install openpyxl",
            )
            return None

        timestamp = datetime.now()
        report_path = os.path.join(self.get_export_base_dir(), self.build_report_filename(timestamp))
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Results"
        worksheet.freeze_panes = "A5"

        thin_side = Side(style="thin", color="B7C9E2")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(italic=True, color="666666")
        section_font = Font(size=11, bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="5B9BD5")
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "THERMO CALCULATOR REPORT"
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = center_alignment
        worksheet.row_dimensions[1].height = 24

        worksheet.merge_cells("A2:D2")
        worksheet["A2"] = f"Exported: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet["A2"].font = subtitle_font
        worksheet["A2"].alignment = center_alignment

        def style_row(row_index, max_col=4, fill=None, font=None):
            for col_index in range(1, max_col + 1):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment
                if fill is not None:
                    cell.fill = fill
                if font is not None:
                    cell.font = font

        def write_section_header(row_index, title):
            worksheet.cell(row=row_index, column=1, value=title)
            style_row(row_index, fill=section_fill, font=section_font)
            return row_index + 1

        def write_table_header(row_index, headers):
            for col_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = table_border
                cell.alignment = left_alignment
            return row_index + 1

        row_index = 4
        row_index = write_section_header(row_index, "Conditions")
        row_index = write_table_header(row_index, ["Setting", "Value"])
        for setting, value in self.get_current_conditions():
            worksheet.cell(row=row_index, column=1, value=setting)
            value_cell = worksheet.cell(row=row_index, column=2, value=value)
            for col_index in range(1, 3):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment
            value_cell.alignment = left_alignment
            row_index += 1

        row_index += 1
        row_index = write_section_header(row_index, "Input Composition")
        row_index = write_table_header(row_index, ["Component", "Mol %", "Wt %"])
        for composition_row in self.get_input_composition_rows():
            worksheet.cell(row=row_index, column=1, value=composition_row["Component"])
            for column_index, key in enumerate(["Mol %", "Wt %"], start=2):
                value = composition_row[key]
                cell = worksheet.cell(row=row_index, column=column_index, value=value if value != "" else None)
                cell.border = table_border
                if isinstance(value, (int, float)):
                    cell.alignment = right_alignment
                    cell.number_format = get_excel_number_format(key)
                else:
                    cell.alignment = left_alignment
            worksheet.cell(row=row_index, column=1).border = table_border
            worksheet.cell(row=row_index, column=1).alignment = left_alignment
            row_index += 1

        warning_rows = self.build_report_warning_rows(self.last_result_data)
        if warning_rows:
            row_index += 1
            row_index = write_section_header(row_index, "Warnings")
            row_index = write_table_header(row_index, ["Warning Type", "Details"])
            for warning_row in warning_rows:
                worksheet.cell(row=row_index, column=1, value=warning_row["Warning Type"])
                worksheet.cell(row=row_index, column=2, value=warning_row["Details"])
                for col_index in range(1, 3):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.border = table_border
                    cell.alignment = left_alignment
                row_index += 1

        row_index += 1
        row_index = write_section_header(row_index, "Results")
        row_index = write_table_header(row_index, ["Property", "Value", "Unit", "Notes"])
        for result_row in self.build_results_rows(self.last_result_data):
            worksheet.cell(row=row_index, column=1, value=result_row["Property"])
            value = result_row["Value"]
            value_cell = worksheet.cell(row=row_index, column=2, value=value if value != "" else None)
            worksheet.cell(row=row_index, column=3, value=result_row["Unit"])
            worksheet.cell(row=row_index, column=4, value=result_row["Notes"])

            for col_index in range(1, 5):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.border = table_border
                cell.alignment = left_alignment

            if isinstance(value, (int, float)):
                value_cell.alignment = right_alignment
                value_cell.number_format = get_excel_number_format(result_row["Unit"])

            row_index += 1

        worksheet.column_dimensions["A"].width = 32
        worksheet.column_dimensions["B"].width = 22
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 52

        try:
            workbook.save(report_path)
        except Exception as exc:
            QMessageBox.critical(self, "Print Results", f"Failed to save Excel report:\n{exc}")
            return None

        QMessageBox.information(self, "Print Results", f"Excel report saved to:\n{report_path}")

        if hasattr(os, "startfile"):
            try:
                os.startfile(report_path)
            except OSError:
                pass

        return report_path

    def setup_flow_tab(self):
        """Initialize the Flow tab controls from the current UI definition."""
        if not hasattr(self.ui, "comboBox_select_units"):
            return

        self.ui.comboBox_select_units.clear()
        self.ui.comboBox_select_units.addItems(FLOW_UNIT_ORDER)
        self.ui.comboBox_select_desired_units.clear()
        self.ui.comboBox_select_desired_units.addItems(FLOW_UNIT_ORDER)
        self.ui.comboBox_select_units.setCurrentText("kg/h")
        self.ui.comboBox_select_desired_units.setCurrentText("t/d")
        self._configure_copyable_flow_output(self.ui.lineEdit_result)
        self._configure_copyable_flow_output(self.ui.lineEdit_in_desired_units)
        self.update_flow_conversion()

    def _configure_copyable_flow_output(self, line_edit):
        """Keep flow outputs read-only while preserving selection and copy behavior."""
        line_edit.setReadOnly(True)
        line_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        line_edit.setContextMenuPolicy(Qt.ContextMenuPolicy.DefaultContextMenu)
        line_edit.setCursor(Qt.CursorShape.IBeamCursor)
        line_edit.setDragEnabled(True)

    def update_flow_conversion(self):
        """Recalculate the Flow-tab conversion using the latest density state."""
        if not hasattr(self.ui, "lineEdit_enter_flow"):
            return

        from_unit = self.ui.comboBox_select_units.currentText().strip()
        to_unit = self.ui.comboBox_select_desired_units.currentText().strip()
        self.ui.lineEdit_in_desired_units.setText(to_unit)

        if not from_unit or not to_unit:
            self.ui.lineEdit_result.clear()
            return

        input_text = self.ui.lineEdit_enter_flow.text()
        if not input_text.strip():
            self.ui.lineEdit_result.clear()
            return

        try:
            value = parse_flow_input(input_text)
        except ValueError:
            self.ui.lineEdit_result.setText("Invalid input")
            return

        if value is None:
            self.ui.lineEdit_result.clear()
            return

        try:
            converted_value = convert_flow(
                value,
                from_unit,
                to_unit,
                density_normal_kg_m3=self.density_normal_kg_m3,
                density_standard_kg_m3=self.density_standard_kg_m3,
            )
        except ValueError as exc:
            self.ui.lineEdit_result.setText(str(exc))
            return

        self.ui.lineEdit_result.setText(format_flow_value(converted_value))

    def populate_comboboxes(self):
        """Populate the ComboBox widgets with options."""
        # Component Selection ComboBox
        self.ui.comboBox_select_components.addItems(list(MOLECULAR_WEIGHTS.keys()))

        # Temperature ComboBox
        cryo_temps = [f"{t} °C" for t in range(-250, 1, 10)]  # -250..0 step 10
        existing_temps = ["0 °C", "5 °C", "10 °C", "15 °C", "25 °C", "50 °C", "100 °C", "150 °C", "200 °C"]
        temps = []
        for t in cryo_temps + existing_temps:
            if t not in temps:
                temps.append(t)
        self.ui.comboBox_select_temperature.addItems(temps)
        self.ui.comboBox_select_temperature.setCurrentText("0 °C") # Default

        # Pressure ComboBox
        self.ui.comboBox_select_pressure.addItems(["1 atm", "5 atm", "10 atm", "20 atm", "50 atm", "100 atm"])
        self.ui.comboBox_select_pressure.setCurrentText("1 atm") # Default

        # EOS ComboBox
        self.ui.comboBox_select_EOS.addItems(["PRMIX"]) # Only PRMIX for now
        # Add more if MixtureCalculator supports them, e.g., ["PRMIX", "SRKMIX"]


    def setup_table(self):
        """Set up the composition table headers and initial row."""
        self.ui.tableWidget.setColumnCount(3)
        self.ui.tableWidget.setHorizontalHeaderLabels(["Component", "Mol %", "Wt %"])
        # Set reasonable column widths (adjust as needed)
        self.ui.tableWidget.setColumnWidth(0, 150)
        self.ui.tableWidget.setColumnWidth(1, 80)
        self.ui.tableWidget.setColumnWidth(2, 80)
        header = self.ui.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive) # Component name resizable
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch) # Stretch % columns
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)

        # Add the non-editable "Total" row
        self.ui.tableWidget.setRowCount(1)
        total_item = QTableWidgetItem("Total")
        total_item.setFlags(Qt.ItemFlag.ItemIsEnabled) # Make non-editable
        self.ui.tableWidget.setItem(0, 0, total_item)
        # Initialize Total value cells
        zero_item1 = QTableWidgetItem("0.00")
        zero_item1.setFlags(Qt.ItemFlag.ItemIsEnabled)
        self.ui.tableWidget.setItem(0, 1, zero_item1)
        zero_item2 = QTableWidgetItem("0.00")
        zero_item2.setFlags(Qt.ItemFlag.ItemIsEnabled)
        self.ui.tableWidget.setItem(0, 2, zero_item2)


    def connect_signals(self):
        """Connect UI element signals to methods."""
        self.ui.comboBox_select_components.currentIndexChanged.connect(self.add_component)
        self.ui.comboBox_select_EOS.currentIndexChanged.connect(self.update_eos)

        # Connect Radio Buttons
        self.ui.radioButton_mol_percent.toggled.connect(self.on_input_basis_changed)
        # Wt% toggled is implicitly handled by Mol% toggling if in a button group,
        # but connecting both is safer if group isn't used or for explicit clarity.
        self.ui.radioButton_wt_percent.toggled.connect(self.on_input_basis_changed)

        # Connect Buttons
        self.ui.remove_component_button.clicked.connect(self.remove_component)
        self.ui.clear_all_button.clicked.connect(self.clear_all)
        self.ui.delete_all_button.clicked.connect(self.clear_all) # Connect the second clear button
        self.ui.go_button.clicked.connect(self.calculate_and_display)
        self.ui.pushButton_normalize.clicked.connect(self.normalize_composition)
        if hasattr(self.ui, "printResultsButton"):
            self.ui.printResultsButton.clicked.connect(self.export_results_to_excel)
        if hasattr(self.ui, "lineEdit_enter_flow"):
            self.ui.lineEdit_enter_flow.textChanged.connect(self.update_flow_conversion)
            self.ui.comboBox_select_units.currentIndexChanged.connect(self.update_flow_conversion)
            self.ui.comboBox_select_desired_units.currentIndexChanged.connect(self.update_flow_conversion)

        # Connect Table Change Signal
        self.ui.tableWidget.itemChanged.connect(self.on_table_item_changed)

        # Optional: Group Radio Buttons programmatically (good practice)
        self.basis_group = QButtonGroup(self) # Create group associated with MainWindow
        self.basis_group.addButton(self.ui.radioButton_mol_percent)
        self.basis_group.addButton(self.ui.radioButton_wt_percent)

        # Step 3: Connect temperature and pressure combo boxes to clear results_list
        self.ui.comboBox_select_temperature.currentIndexChanged.connect(self.on_temperature_or_pressure_changed)
        self.ui.comboBox_select_pressure.currentIndexChanged.connect(self.on_temperature_or_pressure_changed)


    def update_eos(self):
        """Update the EOS in the calculator when the ComboBox changes."""
        # Step 5: Reset progress bar to 0 on EOS change
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        self.invalidate_results()
        selected_eos = self.ui.comboBox_select_EOS.currentText()
        self.calculator.set_eos(selected_eos)

    def add_component(self):
        """Add the selected component to the list and table."""
        # Step 5: Reset progress bar to 0 on add component
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        component = self.ui.comboBox_select_components.currentText()

        # Don't add the dummy entry or duplicates
        if not component.strip():
            return # Ignore the dummy empty selection

        # Check for duplicates in the list_widget (selected_components_list)
        current_items = [self.ui.selected_components_list.item(i).text() for i in range(self.ui.selected_components_list.count())]
        if component in current_items:
            # Highlight combo red and disable Go (or show message box)
            palette = self.ui.comboBox_select_components.palette()
            palette.setColor(QPalette.ColorRole.Base, QColor('red'))
            self.ui.comboBox_select_components.setPalette(palette)
            # self.ui.go_button.setEnabled(False) # Disabling Go might be too strict
            QMessageBox.warning(self, "Duplicate", f"Component '{component}' is already selected.")
            # Reset combo index to dummy after showing message
            self.ui.comboBox_select_components.setCurrentIndex(0)
            return

        # Reset combo box color if previously red
        palette = self.ui.comboBox_select_components.palette()
        palette.setColor(QPalette.ColorRole.Base, QColor('white'))
        self.ui.comboBox_select_components.setPalette(palette)

        # Add to the list_widget
        self.ui.selected_components_list.addItem(component)

        # Add to the tableWidget
        total_row_index = self.ui.tableWidget.rowCount() - 1
        self.ui.tableWidget.insertRow(total_row_index)
        # Put the component name in column 0 (non-editable)
        name_item = QTableWidgetItem(component)
        name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable) # Make name non-editable
        self.ui.tableWidget.setItem(total_row_index, 0, name_item)

        # Add empty items for Mol% and Wt% columns
        mol_item = QTableWidgetItem("") # Start empty
        wt_item = QTableWidgetItem("") # Start empty
        self.ui.tableWidget.setItem(total_row_index, 1, mol_item)
        self.ui.tableWidget.setItem(total_row_index, 2, wt_item)

        # Re-apply basis logic to set correct editability/colors for the new row
        self.on_input_basis_changed() # This will handle flags and background

        # Recalculate totals
        self.update_table_total()

        # Reset combo index to dummy/placeholder
        self.ui.comboBox_select_components.setCurrentIndex(0)

        # Clear results after successful add because prior calculations are now stale.
        self.invalidate_results()

    def remove_component(self):
        """Remove the selected component from the list and table."""
        # Step 5: Reset progress bar to 0 on remove component
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        selected_list_items = self.ui.selected_components_list.selectedItems()

        if not selected_list_items:
            # If nothing selected, maybe remove the last one added? Or show message.
            if self.ui.selected_components_list.count() > 0:
                 current_row = self.ui.selected_components_list.currentRow()
                 if current_row < 0: # If nothing is actively selected, use last item
                      current_row = self.ui.selected_components_list.count() - 1
                 item_to_remove = self.ui.selected_components_list.takeItem(current_row) # Remove from list
                 if item_to_remove:
                      self._remove_component_from_table(item_to_remove.text()) # Remove from table
            else:
                 QMessageBox.information(self, "Remove", "No components to remove.")
                 return
        else:
            # Remove all selected items
            for item in selected_list_items:
                row = self.ui.selected_components_list.row(item)
                comp_name = item.text()
                self.ui.selected_components_list.takeItem(row) # Remove from list
                self._remove_component_from_table(comp_name) # Remove from table by name

        self.update_table_total()
        # Reset combo color if it was red
        if self.ui.selected_components_list.count() == 0:
            palette = self.ui.comboBox_select_components.palette()
            palette.setColor(QPalette.ColorRole.Base, QColor('white'))
            self.ui.comboBox_select_components.setPalette(palette)

        # Clear results after successful remove because prior calculations are now stale.
        self.invalidate_results()

    def _remove_component_from_table(self, comp_name):
        """Helper to remove the row matching comp_name from the table."""
        # Iterate backwards to avoid index issues when removing rows
        for row_index in range(self.ui.tableWidget.rowCount() - 2, -1, -1): # Stop before Total row
            table_item = self.ui.tableWidget.item(row_index, 0)
            if table_item and table_item.text() == comp_name:
                self.ui.tableWidget.removeRow(row_index)
                break # Assume component names are unique

    def clear_all(self):
        """Clear all inputs, selections, and results."""
        # Step 5: Reset progress bar to 0 on clear all
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        # Clear results at the start because all prior calculations are now stale.
        self.invalidate_results()
        # Clear list
        self.ui.selected_components_list.clear()

        # Clear table (remove all rows except the header and 'Total' row)
        # Iterate backwards from second-to-last row up to first row
        while self.ui.tableWidget.rowCount() > 1:
             self.ui.tableWidget.removeRow(0) # Remove the first component row repeatedly

        # Clear results list
        # self.ui.results_list.clear() # Already cleared above

        # Reset calculator
        self.calculator.set_components({})

        # Reset UI states
        self.ui.go_button.setEnabled(True) # Re-enable Go button
        palette = self.ui.comboBox_select_components.palette()
        palette.setColor(QPalette.ColorRole.Base, QColor('white'))
        self.ui.comboBox_select_components.setPalette(palette)
        self.ui.comboBox_select_components.setCurrentIndex(0) # Reset selection
        # Reset T/P/EOS combos to defaults maybe?
        self.ui.comboBox_select_temperature.setCurrentText("0 °C")
        self.ui.comboBox_select_pressure.setCurrentText("1 atm")
        self.ui.comboBox_select_EOS.setCurrentIndex(0)

        # Reset basis to Mol %
        self.ui.radioButton_mol_percent.setChecked(True)
        # on_input_basis_changed will be triggered by setChecked, which calls update_table_total

        # Explicitly update total fields after clearing
        self.update_table_total()

    def on_input_basis_changed(self):
        """Update table column appearance and editability based on radio buttons."""
        is_mol_basis = self.ui.radioButton_mol_percent.isChecked()
        active_col = 1 if is_mol_basis else 2
        inactive_col = 2 if is_mol_basis else 1

        # Block signals temporarily to prevent itemChanged triggers during modification
        self.ui.tableWidget.blockSignals(True)

        row_count = self.ui.tableWidget.rowCount()
        total_row = row_count - 1 # Index of the 'Total' row

        for row in range(row_count):
            active_item = self.ui.tableWidget.item(row, active_col)
            inactive_item = self.ui.tableWidget.item(row, inactive_col)

            # Ensure items exist (might not for newly added row before full setup)
            if active_item is None:
                active_item = QTableWidgetItem("")
                self.ui.tableWidget.setItem(row, active_col, active_item)
            if inactive_item is None:
                inactive_item = QTableWidgetItem("")
                self.ui.tableWidget.setItem(row, inactive_col, inactive_item)

            # Set background colors
            active_item.setBackground(QColor("white"))
            inactive_item.setBackground(QColor("lightGray"))

            # Set flags (editability)
            if row != total_row: # Component rows
                active_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable)
                inactive_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # Not editable
            else: # Total row
                active_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # Not editable
                inactive_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # Not editable

        # Unblock signals
        self.ui.tableWidget.blockSignals(False)
        # Recalculate derived (inactive) column values and then totals for the active column
        self._recalculate_inactive_column()
        self.invalidate_results()
        self.update_table_total() # Update totals after changing basis

    def on_table_item_changed(self, item):
        """Recalculate total when a percentage value changes."""
        # Step 5: Reset progress bar to 0 on composition change
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        # Only update if the changed item is in an active percentage column and not the total row
        active_col = 1 if self.ui.radioButton_mol_percent.isChecked() else 2
        if item.column() == active_col and item.row() != self.ui.tableWidget.rowCount() - 1:
             self.invalidate_results()
             # Basic validation: try converting to float, set red background if invalid?
             try:
                 float(item.text())
                 item.setBackground(QColor("white")) # Reset background on valid input
             except ValueError:
                 if item.text().strip(): # Don't color empty cells red
                      item.setBackground(QColor("pink")) # Indicate invalid number format
             # Live update of the derived inactive column
             self._recalculate_inactive_column()
             self.update_table_total()

    def update_table_total(self):
        """Sum the percentages in the active column and update the Total cell."""
        active_col = 1 if self.ui.radioButton_mol_percent.isChecked() else 2
        row_count = self.ui.tableWidget.rowCount()
        total_row_index = row_count - 1
        total_percent = 0.0
        valid_input = True

        # Iterate only component rows (up to total_row_index)
        for r in range(total_row_index):
            percent_item = self.ui.tableWidget.item(r, active_col)
            if percent_item:
                try:
                    value = float(percent_item.text()) if percent_item.text().strip() else 0.0
                    total_percent += value
                except ValueError:
                     # If text isn't a valid float, treat as error for sum validation
                     if percent_item.text().strip(): # Only count as error if not empty
                          valid_input = False
                     # Don't add to total, let the 100% check fail

        # Update the total cell in the active column
        total_item_active = self.ui.tableWidget.item(total_row_index, active_col)
        if not total_item_active:
            total_item_active = QTableWidgetItem()
            self.ui.tableWidget.setItem(total_row_index, active_col, total_item_active)
        total_item_active.setText(f"{total_percent:.4f}")
        total_item_active.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # Ensure non-editable

        # Clear the total cell in the inactive column
        inactive_col = 2 if active_col == 1 else 1
        total_item_inactive = self.ui.tableWidget.item(total_row_index, inactive_col)
        if not total_item_inactive:
             total_item_inactive = QTableWidgetItem()
             self.ui.tableWidget.setItem(total_row_index, inactive_col, total_item_inactive)
        total_item_inactive.setText("") # Clear inactive total
        total_item_inactive.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)

        # Check if total is 100% and input is valid
        is_total_ok = abs(total_percent - 100.0) < 1e-4 # Use tolerance for float comparison
        if is_total_ok and valid_input:
            total_item_active.setForeground(QColor('black'))
            self.ui.go_button.setEnabled(True)
        else:
            total_item_active.setForeground(QColor('red'))
            self.ui.go_button.setEnabled(False)

    def calculate_and_display(self):
        """Gather inputs, run calculations, and display results."""
        # Step 3: Set progress bar to 0 at the start of calculation
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        # Start animating progress bar to 80% over 1 second as soon as Go! is clicked
        self.animate_progress_to(80, 1000)
        self.invalidate_results()

        # --- Gather inputs for worker ---
        comp_names = []
        mol_percents = []
        wt_percents = []
        row_count = self.ui.tableWidget.rowCount() - 1 # Exclude total row
        valid_input = True
        for r in range(row_count):
            comp_item = self.ui.tableWidget.item(r, 0)
            mol_item = self.ui.tableWidget.item(r, 1)
            wt_item = self.ui.tableWidget.item(r, 2)
            if not comp_item: continue
            comp_names.append(comp_item.text())
            try:
                mol_val = float(mol_item.text()) if mol_item and mol_item.text().strip() else 0.0
                mol_percents.append(mol_val)
            except ValueError:
                mol_percents.append(0.0)
                if self.ui.radioButton_mol_percent.isChecked(): valid_input = False
            try:
                wt_val = float(wt_item.text()) if wt_item and wt_item.text().strip() else 0.0
                wt_percents.append(wt_val)
            except ValueError:
                wt_percents.append(0.0)
                if self.ui.radioButton_wt_percent.isChecked(): valid_input = False

        if not valid_input:
            self.ui.results_list.addItem("Error: Invalid numeric input in composition table.")
            self.animate_progress_to(100, 500)
            return
        if not comp_names:
            self.ui.results_list.addItem("Error: No components selected.")
            self.animate_progress_to(100, 500)
            return

        basis = "Mol %" if self.ui.radioButton_mol_percent.isChecked() else "Wt %"
        try:
            temp_text = self.ui.comboBox_select_temperature.currentText()
            T_c = float(temp_text.split("°")[0])
            T_k = T_c + 273.15
        except ValueError:
            self.ui.results_list.addItem("Error: Invalid Temperature selection.")
            self.animate_progress_to(100, 500)
            return
        try:
            pressure_text = self.ui.comboBox_select_pressure.currentText()
            pressure_atm = float(pressure_text.split(" ")[0])
            pressure_pa = pressure_atm * 101325.0
            if pressure_pa <= 0: raise ValueError("Pressure must be positive")
        except ValueError:
            self.ui.results_list.addItem("Error: Invalid Pressure selection.")
            self.animate_progress_to(100, 500)
            return

        # --- Threaded calculation setup ---
        self.ui.go_button.setEnabled(False)
        if hasattr(self.ui, "printResultsButton"):
            self.ui.printResultsButton.setEnabled(False)
        self.worker_thread = QThread()
        self.worker = CalculationWorker(
            calculator=self.calculator,
            lhv_data=self.lhv_database,
            comp_names=comp_names,
            mol_percents=mol_percents,
            wt_percents=wt_percents,
            basis=basis,
            T_k=T_k,
            pressure_pa=pressure_pa,
            pressure_atm=pressure_atm
        )
        self.worker.moveToThread(self.worker_thread)
        self.worker.result.connect(self.on_calculation_result)
        self.worker.error.connect(self.on_calculation_error)
        self.worker.finished.connect(self.on_calculation_finished)
        self.worker_thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker_thread.start()
        return

    def on_calculation_result(self, result_data):
        # Update the UI with calculation results (runs in main thread)
        self.last_result_data = dict(result_data)
        self.density_actual_kg_m3 = result_data.get("density_actual_kg_m3")
        self.density_normal_kg_m3 = result_data.get("density_normal_kg_m3")
        self.density_standard_kg_m3 = result_data.get("density_standard_kg_m3")
        self.set_thermo_warning_messages(result_data.get("warnings") or [])
        self.update_flow_conversion()
        self.ui.results_list.clear()
        self.ui.results_list.addItem(f"Calculating for {result_data['basis']} basis...")
        self.ui.results_list.addItem(f"Model / EOS: {result_data.get('model_display', self.calculator.eos)}")
        self.ui.results_list.addItem(f"Avg. Molecular Wt: {result_data['mw']:.2f} g/mol")
        # Density/Phase
        density_result = result_data['density_result']
        phase = result_data['phase']
        error = result_data['density_error']
        if error:
            self.ui.results_list.addItem("Phase @ selected conditions: N.A.")
            self.ui.results_list.addItem("Density @ selected conditions: N.A.")
            self.ui.results_list.addItem(f"Density/Phase Error: {error}")
        elif phase == "Two-Phase" and density_result is not None:
            if isinstance(density_result, tuple) and len(density_result) == 2:
                density_liq, density_gas = density_result
                self.ui.results_list.addItem(f"Phase @ selected conditions: {phase}")
                if density_liq is not None:
                    self.ui.results_list.addItem(f"  Density @ selected conditions (Liq): {density_liq:.3f} kg/m³")
                else:
                    self.ui.results_list.addItem("  Density @ selected conditions (Liq): N.A.")
                if density_gas is not None:
                    self.ui.results_list.addItem(f"  Density @ selected conditions (Vap): {density_gas:.3f} kg/m³")
                else:
                    self.ui.results_list.addItem("  Density @ selected conditions (Vap): N.A.")
            else:
                self.ui.results_list.addItem(f"Phase @ selected conditions: {phase} (Result format unexpected)")
        elif density_result is not None:
            self.ui.results_list.addItem(f"Phase @ selected conditions: {phase or 'Could not determine.'}")
            self.ui.results_list.addItem(f"Density @ selected conditions: {density_result:.3f} kg/m³")
        elif phase:
            self.ui.results_list.addItem(f"Phase @ selected conditions: {phase}")
            self.ui.results_list.addItem("Density @ selected conditions: N.A.")
        else:
            self.ui.results_list.addItem("Phase @ selected conditions: Could not determine.")
            self.ui.results_list.addItem("Density @ selected conditions: N.A.")

        def add_reference_density_line(label, density_value, phase_value, error_value):
            if density_value is not None:
                self.ui.results_list.addItem(f"{label}: {density_value:.3f} kg/m³")
                return
            detail = ""
            if error_value:
                detail = f" ({error_value})"
            elif phase_value == "Two-Phase":
                detail = " (Two-Phase)"
            self.ui.results_list.addItem(f"{label}: N.A.{detail}")

        add_reference_density_line(
            "Density @ normal conditions",
            result_data.get("density_normal_kg_m3"),
            result_data.get("density_normal_phase"),
            result_data.get("density_normal_error"),
        )
        add_reference_density_line(
            "Density @ standard conditions",
            result_data.get("density_standard_kg_m3"),
            result_data.get("density_standard_phase"),
            result_data.get("density_standard_error"),
        )
        # Bubble Point
        bubble_point = result_data['bubble_point']
        bp_error = result_data['bp_error']
        pressure_atm = result_data['pressure_atm']
        if bp_error:
            self.ui.results_list.addItem(f"Bubble Point Error: {bp_error}")
        elif bubble_point is not None:
            self.ui.results_list.addItem(f"Bubble Point @ {pressure_atm} atm: {bubble_point:.2f} °C")
        else:
            self.ui.results_list.addItem(f"Bubble Point: Calculation failed.")
        # LHV
        if self.lhv_data_loaded:
            mixture_lhv = result_data['mixture_lhv']
            missing_lhv = result_data['missing_lhv']
            if missing_lhv:
                self.ui.results_list.addItem(f"LHV Warning: No data for: {', '.join(missing_lhv)}")
            self.ui.results_list.addItem("-" * 40)
            lhv_display_values = build_lhv_display_values(mixture_lhv, result_data['mw'])
            volumetric_lines = [
                ("Mixture LHV = ", "MJ/Nm³"),
                ("= ", "kcal/Nm³"),
                ("= ", "MMkcal/Nm³"),
                ("= ", "GJ/Nm³"),
                ("= ", "MMBtu/Nm³"),
            ]
            for prefix, unit in volumetric_lines:
                self.ui.results_list.addItem(
                    f"{prefix}{format_lhv_display_value(lhv_display_values['volumetric'][unit])} {unit}"
                )

            self.ui.results_list.addItem("")
            self.ui.results_list.addItem("Mass basis:")
            for unit in ["MJ/kg", "MJ/t", "GJ/kg", "GJ/t", "kcal/kg", "kcal/t", "MMkcal/kg", "MMkcal/t", "MMBtu/kg", "MMBtu/t"]:
                self.ui.results_list.addItem(
                    f"= {format_lhv_display_value(lhv_display_values['mass_basis'][unit])} {unit}"
                )
        else:
            self.ui.results_list.addItem("-" * 40)
            self.ui.results_list.addItem("Mixture LHV = N/A (DB not loaded)")
        self.animate_progress_to(100, 500)

    def on_calculation_error(self, error_msg):
        self.invalidate_results(clear_visible_results=False)
        self.ui.results_list.clear()
        self.ui.results_list.addItem(error_msg)
        self.animate_progress_to(100, 500)

    def on_calculation_finished(self):
        self.ui.go_button.setEnabled(True)
        if hasattr(self.ui, "printResultsButton"):
            self.ui.printResultsButton.setEnabled(True)

    def normalize_composition(self):
        """Normalize the active composition column (mol% or wt%) so the sum is 100%."""
        # Determine active column
        active_col = 1 if self.ui.radioButton_mol_percent.isChecked() else 2
        row_count = self.ui.tableWidget.rowCount()
        total_row = row_count - 1

        # Gather values from the active column (excluding total row)
        values = []
        for row in range(total_row):
            item = self.ui.tableWidget.item(row, active_col)
            try:
                val = float(item.text()) if item and item.text().strip() else 0.0
            except ValueError:
                val = 0.0
            values.append(val)

        total = sum(values)
        if total == 0:
            QMessageBox.warning(self, "Normalize", "Cannot normalize: total is zero.")
            return

        # Block signals to avoid recursion during update
        self.ui.tableWidget.blockSignals(True)
        norm_values = []
        # Normalize and round all but the last value
        for row, val in enumerate(values):
            if row < total_row - 1:
                norm_val = round((val / total) * 100, 4)
                norm_values.append(norm_val)
            else:
                # Last value: set to 100 - sum of others, rounded to 4 decimals
                last_val = round(100.0 - sum(norm_values), 4)
                norm_values.append(last_val)
        # Set the normalized values in the table
        for row, norm_val in enumerate(norm_values):
            item = self.ui.tableWidget.item(row, active_col)
            if item:
                item.setText(f"{norm_val:.4f}")
        self.ui.tableWidget.blockSignals(False)

        self.invalidate_results()
        self.update_table_total()
        # Optionally, provide user feedback in the results list
        if hasattr(self.ui, 'results_list'):
            self.ui.results_list.addItem("Composition normalized to 100%.")

    def on_temperature_or_pressure_changed(self):
        """Clear the results_list when temperature or pressure is changed."""
        # Step 5: Reset progress bar to 0 on T/P change
        if hasattr(self.ui, 'progressBar'):
            self.ui.progressBar.setValue(0)
        self.invalidate_results()

    def clear_results_list(self):
        """Clear the results_list widget."""
        self.ui.results_list.clear()

    def animate_progress_to(self, target_value, duration_ms=500):
        """Animate the progress bar to the target value over the given duration (ms)."""
        if not hasattr(self.ui, 'progressBar'):
            return
        # Stop any existing timer
        if self.progress_timer is not None:
            self.progress_timer.stop()
            self.progress_timer.deleteLater()
            self.progress_timer = None
        # Set up animation state
        self.progress_target = target_value
        self.progress_animation_duration = duration_ms
        self.progress_animation_start_value = self.ui.progressBar.value()
        self.progress_animation_start_time = QTimer().remainingTime()  # Not used, will use QElapsedTimer below

        # Use QElapsedTimer to track elapsed time
        from PyQt6.QtCore import QElapsedTimer
        self._progress_animation_timer = QElapsedTimer()
        self._progress_animation_timer.start()

        def update_progress():
            elapsed = self._progress_animation_timer.elapsed()
            if elapsed >= self.progress_animation_duration:
                self.ui.progressBar.setValue(self.progress_target)
                if self.progress_timer is not None:
                    self.progress_timer.stop()
                    self.progress_timer.deleteLater()
                    self.progress_timer = None
                return
            # Linear interpolation
            start = self.progress_animation_start_value
            end = self.progress_target
            value = start + (end - start) * (elapsed / self.progress_animation_duration)
            self.ui.progressBar.setValue(int(value))

        self.progress_timer = QTimer(self)
        self.progress_timer.timeout.connect(update_progress)
        self.progress_timer.start(20)


# --- Main Execution Block (Modified) ---
def main():
    app = QApplication(sys.argv)
    # It's good practice to set application name and version if distributing
    # app.setApplicationName("ThermoCalculator")
    # app.setApplicationVersion("1.1")

    app.setStyle(QStyleFactory.create("Windows"))
    app.setPalette(app.style().standardPalette())

    #app.setStyle(QStyleFactory.create("WindowsVista"))

    # Load the LHV data before creating the window
    lhv_data_for_app = load_lhv_data(resource_path('lhv_data.db'))

    # Create and show the main window, passing LHV data
    window = MainWindow(lhv_data=lhv_data_for_app)
    window.show()

    sys.exit(app.exec())

if __name__ == "__main__":
    main()
